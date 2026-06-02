import os
import sys

try:
    import openpyxl
except ImportError:
    print("openpyxl not found. Attempting install...")
    os.system("pip3 install openpyxl")
    import openpyxl

def print_xlsx_summary(filepath):
    if not os.path.exists(filepath):
        print(f"Error: {filepath} does not exist")
        return
        
    print(f"\n=========================================")
    print(f"File: {os.path.basename(filepath)}")
    print(f"=========================================")
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        print("Sheets:", wb.sheetnames)
        for sheetname in wb.sheetnames:
            sheet = wb[sheetname]
            print(f"\n--- Sheet: {sheetname} (Sample rows) ---")
            
            # Print first 50 rows
            non_empty_count = 0
            for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                if non_empty_count >= 50 or r_idx >= 200:
                    break
                if any(cell is not None for cell in row):
                    non_empty_count += 1
                    row_str = " | ".join(str(cell)[:40].replace('\n', ' ').strip() if cell is not None else "" for cell in row[:15])
                    print(f"Row {r_idx+1}: {row_str}")
    except Exception as e:
        print(f"Error reading {filepath}: {e}")

print_xlsx_summary("/Users/bongpark/연금시뮬레이터/26년 지급가상계산.xlsx")
print_xlsx_summary("/Users/bongpark/연금시뮬레이터/2024년 연금수지 시뮬레이션 copy.xlsx")
