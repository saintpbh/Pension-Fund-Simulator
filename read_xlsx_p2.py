import os
import openpyxl

filepath = "/Users/bongpark/연금시뮬레이터/26년 지급가상계산.xlsx"
if os.path.exists(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    print("Sheets:", wb.sheetnames)
    for sheetname in wb.sheetnames[:3]: # 상위 3개 시트만
        sheet = wb[sheetname]
        print(f"\n--- Sheet: {sheetname} (Sample rows) ---")
        non_empty = 0
        for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if r_idx >= 100 or non_empty >= 40:
                break
            if any(cell is not None for cell in row):
                non_empty += 1
                row_str = " | ".join(str(cell)[:40].replace('\n', ' ').strip() if cell is not None else "" for cell in row[:15])
                print(f"Row {r_idx+1}: {row_str}")
else:
    print("File not found")
